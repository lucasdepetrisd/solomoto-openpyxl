import easygui as g
import outputs
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from conv_formats import xls_to_xlsx, xlsx_to_xls
import time
import numpy as np
start_time = time.time()

# PROGRAMA QUE OBTIENE LOS CODIGOS DE DOS ARCHIVOS
# LOS COMPARA Y REEMPLAZA EL PRECIO DEL ARCHIVO ORIGEN EN EL DE DESTINO.
# UTILIZA INTERFAZ GRÁFICA BASADA EN TKINTER.
# POSIBLE TODO: ERRORES Y ACTUALIZADOS LOG EN XLSX

# DEFINITIONS

def copy_price(wb, precio, row, column):
    # print(prod)
    # print("Se copia precio: " + precio + " a celda: " + str(column) + ":" + str(row))
    wb.cell(row, column).value = precio

def let_to_num(letter: str):
    number = column_index_from_string(letter)
    return number

# INITIALIZATION

title = "Solomoto"

output1 = g.enterbox("Bienvenido. Primero ingresa el nombre del archivo a crear.", title, "nuevodux")

dir = g.diropenbox( "Elige la carpeta de destino" )

filename1 = g.fileopenbox( 'Elige la lista de precios nuevos' )

text = "Ingresa los siguientes datos:"
input_list = ["Columna Codigo", "Columna Precio", "Fila Inicial", "Fila Final", "Nro Hoja"]
default_list = ["A", "D", "1", "5000", "1"]
output = g.multenterbox(text, title, input_list, default_list)

filename2 = g.fileopenbox( 'Elige la plantilla de Dux' )

c = xls_to_xlsx(filename1, int(output[4]))
d = xls_to_xlsx(filename2, 1)

wc = c.active
wd = d.active

cat_fila_ini = int(output[2])
cat_fila_fin = int(output[3])
cat_col_cod = let_to_num(output[0])
cat_col_prec = let_to_num(output[1])

dux_fila_ini = 4
dux_col_cod = let_to_num('A')
dux_col_bar = let_to_num('B')
dux_col_prec = let_to_num('E')

cat_nro_filas = cat_fila_fin - cat_fila_ini + 1
dux_nro_filas = wd.max_row
print("Nro filas en " + filename1 + ": " + str(cat_nro_filas))
print("Nro filas en " + filename2 + ": " + str(dux_nro_filas))

# rangoCat = catColCod + str(catFilaIni) + ":" + catColCod + str(catFilaFin)
# codc_list = []
# for row in wc[rangoCat]:
#     codc_list.append([cell.value for cell in row])
        
cat_list = []
dux_list = []

# EXECUTION

# CREATE CATALOG CODES LIST

for x in range(cat_fila_ini, cat_fila_fin + 1):
    codprod = wc.cell(x, cat_col_cod).value
    precio = wc.cell(x, cat_col_prec).value
    cat_list.append([codprod, precio])

cat_np = np.array(cat_list)
print("Lista 1 creada.")

# CREATE DUX LIST

for x in range(dux_fila_ini, dux_nro_filas + 1):
    codprod = wd.cell(x, dux_col_cod).value
    codbarra = wd.cell(x, dux_col_bar).value
    dux_list.append([codprod, codbarra])

dux_np = np.array(dux_list)
print("Lista 2 creada.")

lencodp_list = len(cat_list)
lencodpt_list = len(dux_list)

# REPLACE CODES

fer = open('errores.txt', 'w')
fok = open('codigosok.txt', 'w')
errorcount = 0
copycount = 0

msg = g.msgbox("Presiona OK para comenzar la ejecución. Esto puede tardar unos minutos.", title)

print("Comparando listas...")

for x, prodc in enumerate(cat_list):
    # for y, prodd in enumerate(dux_list):
        
    #     if (prodc[0] == prodd[1]) & (prodd[1] != ""):
    #         copy_price(wd, prodc[1], y + dux_fila_ini, dux_col_prec)
    #         fok.write("SE ACTUALIZA \"%s\" " % prodc[0] + "CON PRECIO $%s " % prodc[1] +  "EN FILA " + str(y + dux_fila_ini) + ".\n")
    #         copycount += 1
    #         break
    #     elif y == len(dux_list)-1:
    #         for z, prodd2 in enumerate(dux_list):
    #             if (prodc[0] == prodd2[0]) & (prodd[0] != ""):
    #                 copy_price(wd, prodc[1], z + dux_fila_ini, dux_col_prec)
    #                 fok.write("SE ACTUALIZA \"%s\" " % prodc[0] + "CON PRECIO $%s " % prodc[1] +  "EN FILA " + str(z + dux_fila_ini) + ".\n")
    #                 copycount += 1
    #                 break
    #         else:
    #             # print(str(prodc))
    #             fer.write("NO SE ENCONTRÓ \"%s\" " % prodc[0] +  "EN FILA " + str(x + 1) + ".\n")
    #             errorcount += 1
    #             break

    if prodc[0] in dux_np:
        index = np.where(prodc[0] == dux_list)

    outputs.progreso(x, lencodp_list)
outputs.fin_progreso()

fok.write("\n--------------------------------\nSe actualizaron %s precios." % copycount)
fer.write("\n--------------------------------\nSe encontraron %s errores." % errorcount)
fer.close()
fok.close()
path = dir + "\\" + output1 + ".xlsx"
print('\nEjecución finalizada.\nGuardando archivo .xlsx...')
d.save(path)
print('\nGuardado exitoso.\nConvirtiendo a .xls...')
end_time = time.time()

message = "Actualización finalizada.\nTiempo total de ejecución: {0:.2f} segundos.\nSe actualizaron {1} precios.\nSe encontraron {2} errores.\nSu archivo se encuentra en {3}.\nPresione OK para guardar en formato .xls y terminar.".format(end_time - start_time, copycount, errorcount, path)
msg = g.msgbox(message, title)

xlsx_to_xls(path)
print('\nGuardado exitoso.')