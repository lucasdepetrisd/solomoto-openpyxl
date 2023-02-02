# Copiar filas de un archivo a otro

import outputs
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

p = load_workbook('../productos1.xlsx')
pt = load_workbook('../prodctotal.xlsx')
wp = p.active
wpt = pt.active

p_row_count = wp.max_row
pt_row_count = wpt.max_row
print("Nro filas en productos1.xlsx: " + str(p_row_count))
print("Nro filas en prodctotal.xlsx: " + str(pt_row_count))
codp_list = []
codpt_list = []


def copy_row(x, y):
    tuplapt = tuple(wpt['A' + str(y+2): 'AE' + str(y+2)])
    for z in range(31):
        valor = tuplapt[0][z].internal_value
        wp.cell(x+2, 18+z).value = valor


def copy_product(prod, pt_row):
    # print(prod)
    # print("Se copia producto: " + prod[0] + " a fila: " + str(pt_row))
    wpt.cell(pt_row, 13).value = prod[2]
    wpt.cell(pt_row, 6).value = prod[3]
    wpt.cell(pt_row, 18).value = prod[4]


def add_product(prod, pt_row):
    # print(prod)
    # print("Se añade producto: " + prod[0] + " a fila: " + str(pt_row))
    wpt.cell(pt_row, 1).value = prod[0]
    wpt.cell(pt_row, 3).value = prod[1]
    wpt.cell(pt_row, 13).value = prod[2]
    wpt.cell(pt_row, 6).value = prod[3]
    wpt.cell(pt_row, 18).value = prod[4]


for x in range(2, p_row_count+1, 1):
    codp = wp.cell(x, 2).value
    detallep = wp.cell(x, 3).value
    costo = wp.cell(x, 4).value
    stock = wp.cell(x, 5).value
    fechstock = wp.cell(x, 6).value
    codp_list.append([codp, detallep, costo, stock, fechstock])
    # print(str(x))
    # print(str(codp))

print("Lista 1 creada.")

for x in range(2, pt_row_count+1, 1):
    codpt = wpt.cell(x, 1).value
    codbarra = wpt.cell(x, 2).value
    codpt_list.append([codpt, codbarra])

print("Lista 2 creada.")

lencodp_list = len(codp_list)
lencodpt_list = len(codpt_list)

fp = open(r'../codigosp.txt', 'w')
b = 0

for x, prod in enumerate(codp_list):
    for y, prodpt in enumerate(codpt_list):

        # if (prod[0] == 'GRI43159' and b == 0):
        #     b = 1
        #     # pt.save('../nuevoprod.xlsx')
        #     wait = input("Press Enter to continue...")
        # if b:
        #     print(prod[0] + ' | ' + prodpt[0])
        
        # if prod[0] == '007001010026' and y == 25889:
        #     wait = input("Pause 1")
        
        if prod[0] == prodpt[0]:
            copy_product(prod, y + 2)
            # outputs.mensaje(prod, y, "add")
            fp.write("(%s), " % prod[0] + ", " + str(y+1))
            break
        elif y == lencodpt_list-1:
            for z, prodpt2 in enumerate(codpt_list):
                # if prod[0] == '007001010026' and z == 3054:
                #     wait = input("Pause 2")

                if prod[0] == prodpt2[1]:
                    copy_product(prod, z + 2)
                    # outputs.mensaje(prod, y, "add")
                    fp.write("(%s), " % prod[0] + ", " + str(z+1))
                    break
            else:
                add_product(prod, lencodpt_list + 2)
                # outputs.mensaje(prod, y, "append")
                codpt_list.append([prod[0], prod[0]])
                print(str(prodpt2))
                lencodpt_list += 1
                fp.write("(%s), " % prod[0] + ", " + str(y+1))
                break
    outputs.progreso(x, lencodp_list)

print('\nEjecución finalizada.')
pt.save('../nuevoprod.xlsx')
