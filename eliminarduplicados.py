# Iterar filas y eliminar duplicados

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def eliminar(a):
	ws.delete_rows(a)
	print("Fila " + str(a) + " eliminada")

wb = load_workbook('null_25285844126313319.xlsx')
ws = wb.active
cell = ws.cell(1,8)
encontrados = 0
filastot = 8108

for x in range(filastot, 1, -1):
	nom1 = ws.cell(x, 2).value
	cat1 = ws.cell(x, 7).value
	cuil1 = ws.cell(x, 8).value
	dom1 = ws.cell(x, 16).value
	cel1 = ws.cell(x, 20).value
	print(str(x))
	print("Progreso: " + "%.2f" % (100-(x*100/filastot)) + "%")
	# print(str(nom) + "\t" + str(cat) + "\t" + str(cuil) + "\t" + str(dom) + "\t" + str(cel))
	for y in range(filastot, 1, -1):
		# print(str(y))
		nom2 = ws.cell(y, 2).value
		cat2 = ws.cell(y, 7).value
		cuil2 = ws.cell(y, 8).value
		dom2 = ws.cell(y, 16).value
		cel2 = ws.cell(y, 20).value
		if (x != y):
			if (nom1 == nom2 and cat1 == cat2 and cuil1 == cuil2 and dom1 == dom2 and cel1 == cel2):
				print("Se encontro duplicado en filas " + str(x) + " y " + str(y))
				eliminar(y)
				encontrados += 1
print("Se encontraron " + str(encontrados) + " duplicados")

wb.save('null_25285844126313319.xlsx')