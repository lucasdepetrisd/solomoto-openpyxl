# Obtener marcas y colocar en columna
import outputs
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

def get_first_word(name: str):
    word = ""
    for c in name:
        if c == " ":
            break
        word += c
    return word    

def list_to_column(wb , li: list, row: int, column: int):
    for offset, value in enumerate(li):
        wb.cell(row=row + offset, column=column).value = value
    return wb

w = load_workbook('../../Tablas/diciembre/nuevoprod.xlsx')
wb = w.active
wb = w['Hoja Nueva']

max_rows = get_maximum_rows(sheet_object=wb)
print("Nro filas en nuevoprod.xlsx: " + str(max_rows))
nombres = []
marcas = []

input("Press Enter to continue...")

for x in range(2, max_rows+1, 1):
    nombre = wb.cell(x, 3).value
    nombres.append(nombre)

# print("ultimo nombre ", nombres)

print("Lista creada.")

input("Press Enter to continue...")

lencodp_list = len(nombres)

for x, nombre in enumerate(nombres):
    marca = get_first_word(nombre)
    mensaje= str.format('Copiando Marca {0}: {1}', x, marca)
    marcas.append(marca)

    outputs.progreso(x, lencodp_list, mensaje)

wb = list_to_column(wb, marcas, 1, 28)

print('\nEjecución finalizada.')
# w.save('./nuevoprod.xlsx')
