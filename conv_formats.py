import xlrd
from openpyxl import Workbook, load_workbook
from pathlib import Path
import win32com.client as win32
import os
import gc

def xls_to_xlsx(src_file_path, sheetnum):
    book_xlsx = Workbook()

    if src_file_path.lower().endswith('.xls'):
        book_xls = xlrd.open_workbook(src_file_path)
        
        sheet_xls = book_xls.sheet_by_index(sheetnum - 1)
        # sheet_names = sheet.name()

        # sheet_names = book_xls.sheet_names()
        # for sheet_index, sheet_name in enumerate(sheet_names):
        #     sheet_xls = book_xls.sheet_by_name(sheet_name)

        sheet_xlsx = book_xlsx.active
        sheet_xlsx.title = sheet_xls.name

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    else:
        book_xlsx = load_workbook(src_file_path)

    return book_xlsx

def xlsx_to_xls(src_file_path):
    
    src_file_path = src_file_path.replace("/", "\\")
    filename = os.path.basename(src_file_path)
    xl = win32.gencache.EnsureDispatch('Excel.Application')
    # wb = xl.Workbooks.Open(src_file_path)
    wb = xl.Workbooks.Add(src_file_path)

    if xl.Workbooks.Count > 0: 
        for wbi in xl.Workbooks:
            if wbi.Name == filename[:-1]: 
                wbi.Close()
    wb.SaveAs(src_file_path[:-1] , FileFormat=56)
    os.remove(src_file_path)
    xl.Quit()
    gc.collect()