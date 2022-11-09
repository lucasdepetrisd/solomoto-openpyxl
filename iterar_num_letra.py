# Iterar celdas por filas y determinar si son numeros o letras

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def eliminar(a):
	ws.delete_rows(a)
	print("Fila " + str(a) + " eliminada")

wb = load_workbook('clientestodosresumen3.xlsx')
ws = wb.active
cell = ws.cell(1,1)

for fila in range(5348, 1, -1):
	cell = ws.cell(fila,1)
	if (cell.value is None):
		# print("Celda vacia")
		eliminar(fila)
	elif (str(cell.value).isnumeric()):
		print("Son numeros")
	else:
		print("Son letras")
		eliminar(fila)