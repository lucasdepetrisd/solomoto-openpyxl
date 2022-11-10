# Copiar filas de un archivo a otro

import outputs
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

p = load_workbook('productos.xlsx')
pt = load_workbook('productostipodeta.xlsx')
wp = p.active
wpt = pt.active

p_row_count = wp.max_row
pt_row_count = wpt.max_row
print("Nro filas en productos.xlsx: " + str(p_row_count))
# print("Nro filas en productostipodeta.xlsx: " + str(pt_row_count))
codp_list = []
codpt_list = []

def copy_row(x, y):
    tuplapt = tuple(wpt['A' + str(y+2): 'AE' + str(y+2)])
    for z in range(31):
        valor = tuplapt[0][z].internal_value
        wp.cell(x+2, 18+z).value = valor
	
for x in range(2, p_row_count+1, 1):
    codp = wp.cell(x, 1).value
    codp_list.append(codp)
    # print(str(x))
    # print(str(codp))

for x in range(2, pt_row_count+1, 1):
    codpt = wpt.cell(x, 1).value
    codpt_list.append(codpt)

lencodp_list = len(codp_list)

for x, codp in enumerate(codp_list):
    for y, codpt in enumerate(codpt_list):
        if (codp == codpt):
            print(codp_list[x] + ", " + str(x))
            copy_row(x, y)
            # break
    outputs.progreso(x, lencodp_list)

p.save('productos.xlsx')