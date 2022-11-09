# Iterar celdas por fila y verificar si esta vacia

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def eliminar(a):
	ws.delete_rows(a)
	print("Fila " + str(a) + " eliminada")

wb = load_workbook('null_25285844126313319.xlsx')
ws = wb.active
cell = ws.cell(1,8)

for fila in range(2, 8122):
	cell = ws.cell(fila, 16)
	# if (cell.value is None or str(cell.value)[0] == " "):
	if (cell.value is None):
		print("Fila " + str(fila) + " vacia")
		cell.value = "A"

wb.save('null_25285844126313319.xlsx')